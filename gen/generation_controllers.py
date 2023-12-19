from .fetch_controllers import *
from PIL import Image, ImageDraw, ImageFont
import io
import os
import requests
from PIL import UnidentifiedImageError
import sys
from zipfile import ZipFile
import openpyxl
from openpyxl.styles import *


def generator(name, classs, event_id, eventname, date, position, cords: dict, fontSize: int):  # changes on gen.py
    log(f'Generating for {name}')
    image = Image.open(event_id + '.png')
    draw = ImageDraw.Draw(image)

    if sys.platform == 'linux' or sys.platform == 'linux2':
        font = ImageFont.truetype('Junicode-Italic.ttf', fontSize)
    else:
        font = ImageFont.truetype('arial', fontSize)
    fontSize -= 3
    draw.text((cords['name'][0], cords['name'][1]-fontSize),
              name, font=font, fill=(0, 0, 0))
    draw.text((cords['class'][0], cords['class'][1]-fontSize),
              classs, font=font, fill=(0, 0, 0))
    draw.text((cords['eventname'][0], cords['eventname'][1]-fontSize),
              eventname, font=font, fill=(0, 0, 0))
    draw.text((cords['date'][0], cords['date'][1]-fontSize),
              date, font=font, fill=(0, 0, 0))
    draw.text((cords['position'][0], cords['position'][1]-fontSize),
              position, font=font, fill=(0, 0, 0))

    im_bytes_arr = io.BytesIO()
    image.save(im_bytes_arr, format='PNG')

    return im_bytes_arr.getvalue()


def getTemplate(event_id: str, template_url: str):
    if '.supabase.co' not in template_url:
        raise ValueError

    res = requests.get(template_url)

    try:
        res.content.decode()
        raise UnidentifiedImageError("Please provide correct template")
    except UnicodeDecodeError:
        pass

    with open(event_id + '.png', 'wb') as file:
        file.write(res.content)


def supagenerate(event_id: str, cords: dict, template_url: str, fontSize: int):
    event = fetchEventDetails(event_id)
    mainStudents = fetchMainStudentsFromEvent(event_id)
    onlyParticipantStudents = fetchAllOnlyParticipants(event_id)
    eventname = event['name']
    eventdate = event['date']

    # generating template in local directory

    getTemplate(event_id, template_url)

    # for winners

    metadatas = list()
    log('STARTED GENERATING CERTIFICATES FOR WINNERS')
    for winner in mainStudents['winner']:
        winnerid = winner['student_id']
        studentDetails = fetchStudentDetailsFromId(winnerid)
        im_bytes = generator(f'{studentDetails["first_name"]} {studentDetails["last_name"]}',
                             studentDetails['class'], event_id, eventname, eventdate, 'Winner', cords, fontSize)
        im_name = studentDetails['first_name'] + \
            studentDetails['last_name'] + studentDetails['class']
        metadatas.append({'name': im_name, 'bytes': im_bytes})
    log('GENERATED CERTIFICATES FOR WINNERS')

    # for runner up

    log('STARTED GENERATING CERTIFICATES FOR RUNNER UPS')

    for runnerup in mainStudents['runnerup']:
        runnerupid = runnerup['student_id']
        studentDetails = fetchStudentDetailsFromId(runnerupid)
        im_bytes = generator(f'{studentDetails["first_name"]} {studentDetails["last_name"]}',
                             studentDetails['class'], event_id, eventname, eventdate, 'First-Runner-Up', cords, fontSize)
        im_name = studentDetails['first_name'] + \
            studentDetails['last_name'] + studentDetails['class']
        metadatas.append({'name': im_name, 'bytes': im_bytes})

    log('GENERATED CERTIFICATES FOR RUNNER UPS')

    # for second runner up
    log('STARTED GENERATING CERTIFICATES FOR SECOND RUNNER UPS')
    for runnerup2 in mainStudents['secondrunnerup']:
        runnerup2id = runnerup2['student_id']
        studentDetails = fetchStudentDetailsFromId(runnerup2id)
        im_bytes = generator(f'{studentDetails["first_name"]} {studentDetails["last_name"]}',
                             studentDetails['class'], event_id, eventname, eventdate, 'Second-Runner-Up', cords, fontSize)
        im_name = studentDetails['first_name'] + \
            studentDetails['last_name'] + studentDetails['class']
        metadatas.append({'name': im_name, 'bytes': im_bytes})

    log('GENERATED CERTIFICATES FOR SECOND RUNNER UPS')

    # for participants
    log('STARTED GENERATING CERTIFICATES FOR PARTICIPANTS')
    for participantid in onlyParticipantStudents:
        studentDetails = fetchStudentDetailsFromId(participantid)
        im_bytes = generator(f'{studentDetails["first_name"]} {studentDetails["last_name"]}',
                             studentDetails['class'], event_id, eventname, eventdate, 'Participant', cords, fontSize)
        im_name = studentDetails['first_name'] + \
            studentDetails['last_name'] + studentDetails['class']
        metadatas.append({'name': im_name, 'bytes': im_bytes})
    log('GENERATED CERTIFICATES FOR PARTICIPANTS')

    return metadatas

# legacy code
# def uploadAllToBucket(eventid: str, metadata_arr):

#     for data in metadata_arr:
#         saveToBucket(eventid, data)

#     # deleting template after generating certificates
#     os.remove(eventid + '.png')
#     supabase.storage.from_('templates').remove([eventid + '.png'])


def clearOutputDir():
    files = os.listdir('output')
    for file in files:
        os.remove('output/' + file)


def zipAndUpload(event_id: str, byte_arr: list):

    clearOutputDir()

    with ZipFile(f'output/{event_id}.zip', 'w') as zip:
        for file in byte_arr:
            filename = f"output/{file['name']}.png"
            with open(filename, 'wb') as f:
                f.write(file['bytes'])
            zip.write(filename)

    zipPath = f'output/{event_id}.zip'
    supabase.storage.from_('certificates').upload(path=f'{event_id}.zip', file=zipPath, file_options={'x-upsert': "true", 'content-type':
                                                                                                      'image/png'})

    supabase.storage.from_('templates').remove(f'{event_id}.png')
    os.remove(f'{event_id}.png')


def generateExcelSheet(event_id: str):
    BRANCHES = ['IF', 'EJ', 'CO', 'CE', 'ME']
    wb = openpyxl.Workbook() if f'{event_id}.xls' not in os.listdir('output'
                                                                    ) else openpyxl.load_workbook(f'output/{event_id}.xls')
    event_details = fetchEventDetails(event_id)
    for branch in BRANCHES:
        sheet = wb.create_sheet(branch)

        tydata = supabase.from_('eventparticipant').select("event_id, group!inner(groupmember!inner(student!inner(first_name, last_name, class, enrollment)))").eq(
            'group.groupmember.student.class', f'TY{branch}').eq('event_id', event_id).execute().data
        sydata = supabase.from_('eventparticipant').select("event_id, group!inner(groupmember!inner(student!inner(first_name, last_name, class, enrollment)))").eq(
            'group.groupmember.student.class', f'SY{branch}').eq('event_id', event_id).execute().data
        fydata = supabase.from_('eventparticipant').select("event_id, group!inner(groupmember!inner(student!inner(first_name, last_name, class, enrollment)))").eq(
            'group.groupmember.student.class', f'FY{branch}').eq('event_id', event_id).execute().data

        branchdata: list = list(itertools.chain(tydata, sydata, fydata))
        bold = Font(bold=True)
        for i in range(int(event_details['team_limit'])):
            cell = sheet.cell(row=1, column=i+1)
            cell.value = f'Member {i+1}'
            cell.font = bold
        cell = sheet.cell(row=1, column=(int(event_details['team_limit']))+1)
        cell.value = 'class'
        cell.font = bold
        for i in range(len(branchdata)):
            count = 1

            for member in branchdata[i]['group']['groupmember']:

                cell = sheet.cell(row=i+2, column=count)
                cell.value = member['student']['first_name'] + \
                    member['student']['last_name']
                count += 1
            cell = sheet.cell(row=i+2, column=event_details['team_limit']+1)
            cell.value = branchdata[i]['group']['groupmember'][0]['student']['class']

        wb.save(f'output/{event_id}.xls')


def uploadSheet(event_id):
    supabase.storage.from_('sheets').upload(f'{event_id}.xls', f'output/{event_id}.xls', {
        'x-upsert': "true"})
